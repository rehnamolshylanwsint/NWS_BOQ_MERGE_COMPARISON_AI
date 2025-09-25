import streamlit as st
import pandas as pd
import re, io
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====== CONFIG ======
HEADER_SEARCH_ROWS = 30
AMOUNT_PATTERN = re.compile(r'(AMOUNT|AMT|AED)', re.IGNORECASE)
QTY_PATTERN    = re.compile(r'(QUANTITY|QTY)', re.IGNORECASE)
RATE_PATTERN   = re.compile(r'(RATE|UNIT RATE)', re.IGNORECASE)

# Colors
GREEN_FILL  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # lowest
RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # highest
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # missing
BLUE_FILL   = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")  # mismatch
HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")  # header

# Borders
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ====== HELPERS ======
def detect_header_row(df0: pd.DataFrame):
    """Find likely header row (returns 0-based row index)."""
    best_row, best_score = None, -1
    for r in range(min(HEADER_SEARCH_ROWS, len(df0))):
        row_vals = df0.iloc[r].astype(str).str.strip().str.upper().tolist()
        score, has_amount = 0, False
        for v in row_vals:
            if "ITEM" in v: score += 1
            if "DESC" in v or "DESCRIPTION" in v: score += 1
            if "RATE" in v: score += 1
            if "UNIT" in v: score += 1
            if "QUANTITY" in v or "QTY" in v: score += 1
            if "AMOUNT" in v or v == "AED" or re.search(r'\bAMT\b', v):
                score += 2
                has_amount = True
        if has_amount and score > best_score:
            best_score, best_row = score, r
    return best_row

def to_number(x):
    """Convert cell value to float if possible"""
    if x is None:
        return None
    try:
        if isinstance(x, str):
            s = re.sub(r'[^\d\.\-]', '', x.strip())
            if not s:
                return None
            return float(s)
        return float(x)
    except:
        return None

def style_worksheet(ws, header_row):
    """
    Apply formatting and freeze header.
    header_row here should be the Excel 1-based header row (e.g. 5),
    and we will freeze the row below (A{header_row+1}) so header stays visible.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == header_row:
                cell.font = Font(bold=True, color="000000")
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto column width
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            try:
                length = len(str(row[0]))
                if length > max_length:
                    max_length = length
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row (make header visible)
    ws.freeze_panes = f"A{header_row+1}"

# ====== APP ======
st.set_page_config(page_title="Tender BoQ Comparison", layout="wide")

# Hide footer & GitHub link
st.markdown(
    """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .viewerBadge_container__1QSob {display: none;}
    </style>
    """,
    unsafe_allow_html=True
)

# Sidebar AI Assistant (with chat history)
st.sidebar.header("🤖 AI Assistant")
if "chat_history" not in st.session_state:
    st.session_state["chat_history"] = []

user_question = st.sidebar.text_area("Ask me anything about BoQ comparison:")
if st.sidebar.button("Get Answer"):
    if user_question.strip():
        answer = ""
        if "missing" in user_question.lower():
            answer = "🟨 Missing values = Yellow cells (no entry from contractor)."
        elif "lowest" in user_question.lower():
            answer = "🟩 Lowest values = Green cells (best/lowest rates or amounts)."
        elif "highest" in user_question.lower():
            answer = "🟥 Highest values = Red cells (worst/highest rates or amounts)."
        elif "mismatch" in user_question.lower():
            answer = "🟦 Mismatch values = Blue cells (when Rate × Qty ≠ Amount)."
        else:
            answer = "This tool merges contractor BoQs, highlights Lowest/Highest/Missing/Mismatch values, and exports formatted Excel."
        st.session_state["chat_history"].append(("You", user_question))
        st.session_state["chat_history"].append(("AI", answer))

# Render chat history in sidebar
for speaker, msg in st.session_state["chat_history"]:
    if speaker == "You":
        st.sidebar.markdown(f"**🧑 {speaker}:** {msg}")
    else:
        st.sidebar.markdown(f"**🤖 {speaker}:** {msg}")

# Welcome page / navigation
if "page" not in st.session_state:
    st.session_state["page"] = "welcome"

if st.session_state["page"] == "welcome":
    try:
        st.image("logo.png", width=120)
    except Exception:
        pass

    st.markdown(
        """
        <h1 style="text-align: center; color: #2E86C1;">NWS International</h1>
        <h3 style="text-align: center;">Tender BoQ Merge & Comparison Tool</h3>
        <p style="text-align: center; font-size:16px;">
        Upload contractor BoQs, merge them, compare rates & amounts,<br>
        and download a formatted Excel with highlights.
        </p>
        """,
        unsafe_allow_html=True
    )
    st.info("💡 Use the AI Assistant in the sidebar for quick guidance.")

    if st.button("🚀 Get Started"):
        st.session_state["page"] = "main"
        st.rerun()

# MAIN PAGE
elif st.session_state["page"] == "main":
    st.sidebar.header("⚙️ Settings")
    take_first_three_only = st.sidebar.checkbox("Compare only first 3 contractor columns", value=False)

    # Legend
    st.sidebar.markdown("### 🎨 Legend")
    st.sidebar.markdown("🟩 **Lowest Value**")
    st.sidebar.markdown("🟥 **Highest Value**")
    st.sidebar.markdown("🟨 **Missing Value**")
    st.sidebar.markdown("🟦 **Mismatch (Rate × Qty ≠ Amount)**")

    uploaded_files = st.file_uploader(
        "📂 Upload contractor Excel files",
        type=["xlsx"],
        accept_multiple_files=True
    )

    if uploaded_files and st.button("🔗 Merge & Compare"):
        all_sheets = {}
        for uf in uploaded_files:
            try:
                xls = pd.ExcelFile(uf)
                for sheet in xls.sheet_names:
                    all_sheets.setdefault(sheet, []).append(uf)
            except Exception as e:
                st.error(f"Could not read file {uf.name}: {e}")

        merged_book, summary = {}, []

        for sheet_name, files in all_sheets.items():
            dfs = []
            for idx, uf in enumerate(files):
                df = pd.read_excel(uf, sheet_name=sheet_name, dtype=object)

                unnamed_cols = [c for c in df.columns if str(c).startswith("Unnamed")]
                for c in unnamed_cols:
                    col_vals = df[c]
                    if col_vals.isna().all():
                        df.drop(columns=[c], inplace=True)
                    else:
                        new_name = uf.name.split('.')[0]
                        base = new_name
                        counter = 1
                        while new_name in df.columns:
                            new_name = f"{base}_{counter}"
                            counter += 1
                        df.rename(columns={c: new_name}, inplace=True)

                df.dropna(axis=1, how='all', inplace=True)

                if idx == 0:
                    dfs.append(df)
                else:
                    if df.shape[1] > 2:
                        df_trim = df.iloc[:, 2:].copy()
                    else:
                        df_trim = df.copy()
                    prefix = uf.name.split('.')[0] + "_"
                    df_trim = df_trim.add_prefix(prefix)
                    dfs.append(df_trim)

            try:
                merged_df = pd.concat(dfs, axis=1, ignore_index=False)
            except Exception:
                for i, d in enumerate(dfs):
                    d.columns = [f"c{i}_{j}" for j in range(len(d.columns))]
                merged_df = pd.concat(dfs, axis=1)

            merged_df = merged_df.loc[:, ~merged_df.columns.astype(str).str.startswith("Unnamed")]
            merged_book[sheet_name] = merged_df

        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            for sheet, df in merged_book.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
        bio.seek(0)
        merged_data = bio.getvalue()

        wb = load_workbook(io.BytesIO(merged_data))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df0 = pd.read_excel(io.BytesIO(merged_data), sheet_name=sheet_name, header=None, dtype=object)
            header_row = detect_header_row(df0)
            if header_row is None:
                continue
            header_excel_row = header_row + 1

            headers = [str(ws.cell(row=header_excel_row, column=c).value or "").strip()
                       for c in range(1, ws.max_column + 1)]

            amount_cols = [i for i, h in enumerate(headers, 1) if AMOUNT_PATTERN.search(h)]
            if not amount_cols:
                continue
            if take_first_three_only and len(amount_cols) > 3:
                amount_cols = amount_cols[:3]

            # find rate & qty cols
            rate_cols = [i for i, h in enumerate(headers, 1) if RATE_PATTERN.search(h)]
            qty_cols  = [i for i, h in enumerate(headers, 1) if QTY_PATTERN.search(h)]

            low_count, high_count, missing_count, mismatch_count = 0, 0, 0, 0
            for r in range(header_excel_row + 1, ws.max_row + 1):
                vals = [to_number(ws.cell(row=r, column=c).value) for c in amount_cols]
                present = [v for v in vals if v is not None]
                if present:
                    mn, mx = min(present), max(present)
                    for j, v in enumerate(vals):
                        cell = ws.cell(row=r, column=amount_cols[j])
                        if v is None:
                            cell.fill = YELLOW_FILL
                            missing_count += 1
                        elif v == mn:
                            cell.fill = GREEN_FILL
                            low_count += 1
                        elif v == mx:
                            cell.fill = RED_FILL
                            high_count += 1

                # ---- mismatch check ----
                if rate_cols and qty_cols and amount_cols:
                    rate   = to_number(ws.cell(row=r, column=rate_cols[0]).value)
                    qty    = to_number(ws.cell(row=r, column=qty_cols[0]).value)
                    amount = to_number(ws.cell(row=r, column=amount_cols[0]).value)

                    if rate is not None and qty is not None and amount is not None:
                        expected = rate * qty
                        if abs(expected - amount) > 1e-6:
                            amt_cell = ws.cell(row=r, column=amount_cols[0])
                            amt_cell.fill = BLUE_FILL
                            mismatch_count += 1

            style_worksheet(ws, header_excel_row)
            summary.append([sheet_name, low_count, high_count, missing_count, mismatch_count])

        out_io = io.BytesIO()
        wb.save(out_io)
        out_io.seek(0)

        # Sidebar summary
        st.sidebar.subheader("📊 Summary")
        df_summary = pd.DataFrame(
            summary,
            columns=["Sheet", "Lowest (Green)", "Highest (Red)", "Missing (Yellow)", "Mismatch (Blue)"]
        )
        st.sidebar.dataframe(df_summary, use_container_width=True)

        # Graphs
        st.subheader("📊 Visual Summary")
        if not df_summary.empty:
            df_melt = df_summary.melt(id_vars="Sheet", var_name="Category", value_name="Count")
            fig = px.bar(df_melt, x="Sheet", y="Count", color="Category", barmode="group", title="Comparison Results per Sheet")
            st.plotly_chart(fig, use_container_width=True)

        st.success("✅ Merge & Comparison complete")

        # Preview sheets
        for sheet, df in merged_book.items():
            st.subheader(f"📑 {sheet}")
            safe_df = df.dropna(axis=1, how='all').fillna("").astype(str)
            st.dataframe(safe_df.head(50), use_container_width=True)

        # Download button
        st.download_button(
            "⬇️ Download Highlighted Excel",
            data=out_io.getvalue(),
            file_name="Tender_BoQ_Comparison_Formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
